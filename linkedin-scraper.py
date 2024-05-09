from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
import time
import os
import pandas as pd
from openpyxl import load_workbook, Workbook

options = Options()
options.add_experimental_option("detach", True)

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
linkedin_url = "https://www.linkedin.com/"


def sign_in(email_address, password ):
    email_field = driver.find_element(By.ID, 'session_key')  
    email_field.send_keys(email_address)

    password_field = driver.find_element(By.ID, 'session_password')  
    password_field.send_keys(password)

    sign_in_button = driver.find_element(By.CSS_SELECTOR, '[data-id="sign-in-form__submit-btn"]')
    sign_in_button.click()

def click_reactions_link():
    reactions_link = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "span.social-details-social-counts__reactions-count"))
    )
    reactions_link.click()

def extract_user_data():
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, '.scaffold-finite-scroll__content'))
    )
    users = driver.find_elements(By.CSS_SELECTOR, '.social-details-reactors-tab-body-list-item')
    user_data = []
    for user in users:
        try:
            user_name = user.find_element(By.CSS_SELECTOR, '.text-view-model').text
            profile_link = user.find_element(By.TAG_NAME, 'a').get_attribute('href')
            if len(user_data) > 0 :
                user_data.append({'Post URL': '','Reacted User Name': user_name,
                              'Reacted User Profile Link': profile_link})
            else:
                user_data.append({'Post URL': post_url ,'Reacted User Name': user_name,
                              'Reacted User Profile Link': profile_link})
        except Exception as e:
            print(f"Failed to extract data for a user: {e}")
    return user_data

def get_total_reactions():
    reaction_count_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".social-details-reactors-tab__icon-container span:nth-of-type(2)"))
    )
    if ',' in str(reaction_count_element.text):
        print(str(reaction_count_element.text))
        total_reactions = int(reaction_count_element.text.replace(',', ''))  
    else:
        print(str(reaction_count_element.text))
        total_reactions = int(reaction_count_element.text.replace('.', ''))  
    print("Total users:", total_reactions)
    return total_reactions

def wait_for_more_than_n_elements(locator, count):
    def _predicate(driver):
        return len(driver.find_elements(*locator)) > count
    return _predicate


def scroll_down_modal(total_reactions, modal_selector, speed=5):
    modal = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, modal_selector))
    )
    wait = WebDriverWait(driver, 10)
    current_elements = 0

    while current_elements < total_reactions:
        elements = driver.find_elements(By.CSS_SELECTOR, ".social-details-reactors-tab-body-list-item")
        if elements:
            driver.execute_script("arguments[0].scrollIntoView();", elements[-1])
            try:
                wait.until(wait_for_more_than_n_elements((By.CSS_SELECTOR, ".social-details-reactors-tab-body-list-item"), len(elements)))
                current_elements = len(elements)
                print("Current users:", current_elements)
                time.sleep(0.5 / speed)  # Adjust sleep to control speed and responsiveness
            except TimeoutException:
                print("No new elements loaded after waiting.")
                return
        else:
            break  


############### DEGISKENLERIN OLDUGU KISIM ##################
#pip install openpyxl
#pip install selenium
#pip install pandas

#postun urlsini buraya yapıştır

post_url = str(input("Enter the Post URL: "))
        
#post_url = "https://www.linkedin.com/feed/update/urn:li:activity:7191343366956048385/"

email_address = "serrasadir@sabanciuniv.edu" #kendi linkedin mailini ve şifreni yaz
password = "Al0homora"

workbook_name = 'Book10.xlsx' #Dosya adını yaz - xlsx uzantılı olsun
                             #dosya zaten varsa içine ekler yoksa yeni dosya yaratır

#############################################################

################## SAVE TO EXCEL PART ########################

def save_data_to_excel(workbook_name, user_data):
    new_df = pd.DataFrame(user_data)

    if not os.path.exists(workbook_name):
        wb = Workbook()
        ws = wb.active
        columns = ['Post Url', 'Reacted User Name', 'Reacted User Profile Link']        
        for i, column in enumerate(columns, start=1):
            ws.cell(row=1, column=i, value=column)
        wb.save(workbook_name)
        new_df.to_excel(workbook_name, index=False)
    else:
        wb = load_workbook(workbook_name)
        existing_df = pd.read_excel(workbook_name)
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        combined_df.to_excel(workbook_name, index=False)

    check_df = pd.read_excel(workbook_name)
    print(check_df.tail())

#############################################################

######################## MAIN PART ##########################

driver.get(linkedin_url)
driver.maximize_window()
sign_in(email_address, password)
time.sleep(1)  
driver.get(post_url)  
click_reactions_link()
total_reactions = get_total_reactions()
loaded_users = 0

scroll_down_modal(total_reactions, ".scaffold-finite-scroll__content")

user_data = extract_user_data()

print("Scrape process is done.")

save_data_to_excel(workbook_name, user_data)
print("User data saved to", workbook_name)

#############################################################

######################## COLORING PART ##########################

#CEE Names sarı işaretli
#TR names yeşil işaretli
#Row boyamak için bu celli runla
from openpyxl.styles import PatternFill

tr_workbook_name = "TR Names Excel (updated).xlsx"
dftr = pd.read_excel(tr_workbook_name, sheet_name="Kısa TR Text to Copy")

cee_workbook_name = "CEE Names Excel .xlsx"
dfcee = pd.read_excel(cee_workbook_name, sheet_name="CEE")
dflinkedin = pd.read_excel(workbook_name)

wb = load_workbook(workbook_name)

try:
    ws = wb['Sheet1']
except KeyError:
    print("falling back to Sheet")
    ws = wb['Sheet']

for i in range(0, dflinkedin.shape[1]-1):
    ws.cell(row = 1, column = i+1).value = dflinkedin.columns[i]

for i, cell in enumerate(dflinkedin['Reacted User Name']):
    names = str(cell).split()
    for ceename in dfcee['Text to copy']:
        for n in range(len(names)-1):
            if ceename == names[n]:
              #print(f"Found match: {names[n]}, {i}")  #debug
              for j in range(1,11):
                  ws.cell(i+2,j).fill = PatternFill(start_color='FFD970', end_color='FFD970', fill_type="solid")
for i, cell in enumerate(dflinkedin['Reacted User Name']):
    names = str(cell).split()  # This line is changed
    for trname in dftr[dftr.columns[2]]:
        for n in range(len(names)-1):
            if trname == names[n].upper():
                print(f"Found match: {names[n]}, {i}")  #debug
                for j in range(1, 11):
                    ws.cell(i + 2, j).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type="solid")
wb.save(workbook_name)
wb.close()

print("Modified df saved to excel file successfully.")  #debug message

#############################################################